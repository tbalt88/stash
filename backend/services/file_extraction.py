"""Best-effort text extraction for uploaded files.

Supported:
- PDFs with embedded text → pypdf (pure Python).
- Plain-text / JSON / XML → UTF-8 decode.
- `.docx`, `.pptx` → zipfile + ElementTree walk (no extra deps).
- `.xlsx` → openpyxl, cells joined into rows.
- Everything else → None.

`extract_text` never raises. Every failure returns None so uploads
always succeed even if the file is corrupt or the libs aren't present.
"""

from __future__ import annotations

import io
import logging
import zipfile
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)

try:
    import pypdf  # type: ignore

    _HAS_PYPDF = True
except ImportError:
    _HAS_PYPDF = False

try:
    from openpyxl import load_workbook  # type: ignore

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


_OOXML_DRAWING_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
_OOXML_WORD_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def is_pdf(content_type: str) -> bool:
    ct = (content_type or "").lower()
    return ct == "application/pdf" or ct.endswith("/pdf")


def _extract_pdf_embedded(content: bytes) -> str:
    if not _HAS_PYPDF:
        return ""
    reader = pypdf.PdfReader(io.BytesIO(content))
    parts: list[str] = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n\n".join(p for p in parts if p).strip()


def _extract_docx(content: bytes) -> str:
    """Walk the OOXML body, paragraph by paragraph. Tables are rendered with
    cells separated by tabs so the column structure survives in plain text."""
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        if "word/document.xml" not in z.namelist():
            return ""
        root = ET.fromstring(z.read("word/document.xml"))
    paragraphs: list[str] = []
    for elem in root.iter():
        if elem.tag == f"{_OOXML_WORD_NS}p":
            text = "".join(t.text or "" for t in elem.iter(f"{_OOXML_WORD_NS}t"))
            if text:
                paragraphs.append(text)
        elif elem.tag == f"{_OOXML_WORD_NS}tbl":
            for row in elem.iter(f"{_OOXML_WORD_NS}tr"):
                cells = []
                for cell in row.iter(f"{_OOXML_WORD_NS}tc"):
                    cell_text = "".join(t.text or "" for t in cell.iter(f"{_OOXML_WORD_NS}t"))
                    cells.append(cell_text)
                paragraphs.append("\t".join(cells))
    return "\n".join(paragraphs).strip()


def _extract_pptx(content: bytes) -> str:
    """Slide-by-slide text. The OOXML `<a:t>` element wraps every text run in
    every shape — that includes titles, body text, table cells, and chart
    labels — so a single pass over all matching nodes captures the readable
    content of the deck."""
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        slide_names = sorted(
            n for n in z.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")
        )
        slides: list[str] = []
        for name in slide_names:
            root = ET.fromstring(z.read(name))
            runs = [t.text or "" for t in root.iter(f"{_OOXML_DRAWING_NS}t")]
            slide_text = " ".join(r for r in runs if r)
            if slide_text.strip():
                slides.append(slide_text)
    return "\n\n".join(slides).strip()


def _extract_xlsx(content: bytes) -> str:
    """Every sheet → "## <sheet>\n<row TSV>\n..." block.

    openpyxl is already a dep for xlsx_ingest; reusing it keeps behaviour
    consistent (same parser → same blind spots → same cells visible).
    """
    if not _HAS_OPENPYXL:
        return ""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheets_out: list[str] = []
        for sheet in wb.worksheets:
            if sheet.sheet_state != "visible":
                continue
            rows_out: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                if not any(cell is not None for cell in row):
                    continue
                rows_out.append("\t".join("" if c is None else str(c) for c in row))
            if rows_out:
                sheets_out.append(f"## {sheet.title}\n" + "\n".join(rows_out))
        return "\n\n".join(sheets_out).strip()
    finally:
        wb.close()


def _sanitize_for_postgres(text: str) -> str:
    """Make `text` safe for a Postgres TEXT column.

    Two hazards seen in the wild from pypdf output:
    - null bytes (0x00) — Postgres rejects them outright.
    - unpaired UTF-16 surrogates (U+D800..U+DFFF) — asyncpg can't encode
      them to UTF-8 on the wire. Happens when a PDF's CMap maps a
      character to one half of a surrogate pair without the other.

    Both are stripped. The encode/decode round-trip with
    `errors="replace"` is cheap and handles any other exotic codepoints
    that slip past the surrogate check.
    """
    if "\x00" in text:
        text = text.replace("\x00", "")
    # Drop unpaired surrogates by round-tripping via UTF-8 with replacement.
    text = text.encode("utf-8", errors="replace").decode("utf-8", errors="replace")
    return text


_OOXML_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": _extract_docx,
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": _extract_pptx,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _extract_xlsx,
}


def extract_text(content: bytes, content_type: str) -> str | None:
    """Return extracted text, or None when extraction is not possible/failed.

    Never raises — caller can trust this to not break the upload path.
    """
    try:
        ct = (content_type or "").lower()

        if is_pdf(ct):
            text = _extract_pdf_embedded(content)
            return _sanitize_for_postgres(text) if text else None

        if ct.startswith("text/") or ct in ("application/json", "application/xml"):
            return _sanitize_for_postgres(content.decode("utf-8", errors="replace"))

        extractor = _OOXML_TYPES.get(ct)
        if extractor is not None:
            text = extractor(content)
            return _sanitize_for_postgres(text) if text else None

        return None
    except Exception as exc:
        logger.warning(
            "file_extraction: extract_text failed content_type=%s exception_type=%s",
            content_type,
            type(exc).__name__,
        )
        return None
