"""Parse an XLSX workbook (bytes) into one Stash table per visible sheet.

Shared between the `/files/{id}/ingest-xlsx` endpoint (local file
uploads) and the Google Drive Sheets importer (Drive export → XLSX
bytes). Reuses the CSV inference + coercion so column types end up
identical whether the data came from a CSV, a Sheet, or a workbook.
"""

from __future__ import annotations

import io
import re
from uuid import UUID

from . import table_service
from .csv_inference import coerce_value, infer_column_type


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _slugify(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", (s or "").strip().lower())
    return re.sub(r"_+", "_", s).strip("_")


async def ingest_xlsx_bytes(
    *,
    workspace_id: UUID,
    user_id: UUID,
    content: bytes,
    base_name: str,
    description_template: str,
) -> list[dict]:
    """Create one Stash table per visible sheet with data. Returns the
    list of created table rows (in workbook order).

    `description_template` is `str.format`-ed with `sheet=<title>` per
    sheet so the caller controls the per-table provenance string.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    created: list[dict] = []
    # Only count visible sheets when deciding whether to disambiguate
    # table names — a single visible sheet next to a hidden one shouldn't
    # get a " — sheet" suffix it doesn't need.
    visible_count = sum(1 for s in wb.worksheets if s.sheet_state == "visible")
    multi_sheet = visible_count > 1
    try:
        for sheet in wb.worksheets:
            if sheet.sheet_state != "visible":
                continue
            rows = [[_cell_to_str(c) for c in r] for r in sheet.iter_rows(values_only=True)]
            while rows and all(c == "" for c in rows[-1]):
                rows.pop()
            if not rows:
                continue
            header = rows[0]
            if not any(h.strip() for h in header):
                continue
            data_rows = rows[1:]
            sample = data_rows[:50]

            columns = []
            seen_ids: set[str] = set()
            for ci, name in enumerate(header):
                samples = [(r[ci] if ci < len(r) else "") for r in sample]
                col_id = _slugify(name) or f"col_{ci}"
                base_id = col_id
                n = 2
                while col_id in seen_ids:
                    col_id = f"{base_id}_{n}"
                    n += 1
                seen_ids.add(col_id)
                columns.append(
                    {
                        "id": col_id,
                        "name": (name or f"col_{ci}").strip() or f"col_{ci}",
                        "type": infer_column_type(samples),
                        "order": ci,
                        "required": False,
                        "default": None,
                        "options": None,
                    }
                )

            table_name = f"{base_name} — {sheet.title}" if multi_sheet else base_name
            # Notion DB import uses random col ids (no _slugify); Sheets/CSV
            # use slugified ones. Mix is fine — id collisions within a single
            # table are what matters, and we de-dupe above.
            table = await table_service.create_table(
                workspace_id=workspace_id,
                name=table_name,
                description=description_template.format(sheet=sheet.title),
                columns=columns,
                created_by=user_id,
            )

            payload = []
            for r in data_rows:
                rec = {}
                for ci, col in enumerate(columns):
                    raw = r[ci] if ci < len(r) else ""
                    rec[col["id"]] = coerce_value(raw, col["type"])
                payload.append(rec)
            if payload:
                await table_service.create_rows_batch(
                    table_id=table["id"], rows_data=payload, created_by=user_id
                )

            refreshed = await table_service.get_table(table["id"])
            created.append(refreshed or table)
    finally:
        wb.close()

    return created
