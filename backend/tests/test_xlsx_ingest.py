"""Verify the XLSX ingest service end-to-end against a real openpyxl
workbook — multi-sheet, type inference, and empty-sheet handling.

We mock `table_service.create_table` / `create_rows_batch` / `get_table`
so this exercises the openpyxl walk + the per-sheet shape without
touching the DB.
"""

import io
from datetime import date, datetime
from unittest.mock import AsyncMock, patch
from uuid import uuid4

import pytest
from openpyxl import Workbook

from backend.services.xlsx_ingest import ingest_xlsx_bytes


def _bytes_from(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def mocked_table_service():
    """Capture create_table/create_rows_batch calls so each test can
    inspect what would land in the DB without standing up the DB."""
    created_tables: list[dict] = []
    inserted_rows: dict[str, list[dict]] = {}

    async def fake_create_table(**kwargs):
        tid = uuid4()
        rec = {
            "id": tid,
            "name": kwargs["name"],
            "description": kwargs["description"],
            "columns": kwargs["columns"],
            "owner_user_id": kwargs["owner_user_id"],
            "created_by": kwargs["created_by"],
        }
        created_tables.append(rec)
        return rec

    async def fake_create_rows_batch(*, table_id, rows_data, created_by):
        inserted_rows.setdefault(str(table_id), []).extend(rows_data)
        return rows_data

    async def fake_get_table(table_id):
        for t in created_tables:
            if t["id"] == table_id:
                return t
        return None

    with (
        patch(
            "backend.services.xlsx_ingest.table_service.create_table",
            AsyncMock(side_effect=fake_create_table),
        ),
        patch(
            "backend.services.xlsx_ingest.table_service.create_rows_batch",
            AsyncMock(side_effect=fake_create_rows_batch),
        ),
        patch(
            "backend.services.xlsx_ingest.table_service.get_table",
            AsyncMock(side_effect=fake_get_table),
        ),
    ):
        yield created_tables, inserted_rows


async def test_single_sheet_named_after_base(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws = wb.active
    ws.title = "Only"
    ws.append(["name", "score"])
    ws.append(["alice", 91])
    ws.append(["bob", 88])

    tables = await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="report",
        description_template="src=test sheet={sheet}",
    )
    assert len(tables) == 1
    # Single-sheet workbooks keep the bare base name (no " — sheet" suffix).
    assert created_tables[0]["name"] == "report"
    # Numeric column inferred correctly despite openpyxl returning floats.
    types = {c["name"]: c["type"] for c in created_tables[0]["columns"]}
    assert types == {"name": "text", "score": "number"}


async def test_multi_sheet_creates_one_table_per_tab(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["region", "amount"])
    ws1.append(["US", 1000])
    ws2 = wb.create_sheet("Costs")
    ws2.append(["category", "spend"])
    ws2.append(["cloud", 250])

    tables = await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="Q4",
        description_template="src=test sheet={sheet}",
    )
    assert len(tables) == 2
    names = [t["name"] for t in created_tables]
    assert names == ["Q4 — Sales", "Q4 — Costs"]
    # Description is templated per-sheet.
    assert created_tables[0]["description"] == "src=test sheet=Sales"
    assert created_tables[1]["description"] == "src=test sheet=Costs"


async def test_hidden_sheets_are_skipped(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.append(["k", "v"])
    ws1.append(["a", "b"])
    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden.append(["should", "not appear"])

    tables = await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="book",
        description_template="{sheet}",
    )
    assert len(tables) == 1
    assert created_tables[0]["name"] == "book"


async def test_empty_sheet_is_skipped_silently(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Empty"
    # No rows at all
    ws2 = wb.create_sheet("Real")
    ws2.append(["x", "y"])
    ws2.append([1, 2])

    tables = await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="wb",
        description_template="{sheet}",
    )
    # Only the non-empty sheet should produce a table.
    assert len(tables) == 1
    assert created_tables[0]["name"] == "wb — Real"


async def test_dates_and_booleans_round_trip(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.append(["when", "done"])
    ws.append([date(2026, 5, 21), True])
    ws.append([datetime(2026, 5, 22, 10, 30), False])

    await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="dates",
        description_template="{sheet}",
    )
    cols = {c["name"]: c["type"] for c in created_tables[0]["columns"]}
    # `when` mixes a date (no T) and a datetime (with T) — promoted to datetime.
    assert cols["when"] == "datetime"
    assert cols["done"] == "boolean"


async def test_id_collision_disambiguates(mocked_table_service):
    created_tables, _ = mocked_table_service
    wb = Workbook()
    ws = wb.active
    ws.title = "Dups"
    # Two header cells that slugify to the same id ("name", "Name!").
    ws.append(["name", "Name!"])
    ws.append(["a", "b"])

    await ingest_xlsx_bytes(
        owner_user_id=uuid4(),
        user_id=uuid4(),
        content=_bytes_from(wb),
        base_name="dup",
        description_template="{sheet}",
    )
    ids = [c["id"] for c in created_tables[0]["columns"]]
    # Second column must get a suffixed id so JSONB writes don't clobber.
    assert ids[0] != ids[1]
    assert ids[0] == "name"
    assert ids[1].startswith("name")
